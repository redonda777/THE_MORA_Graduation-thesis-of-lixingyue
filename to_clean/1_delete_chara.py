"""
删除 Excel 中指定字符并输出新文件。

用法示例：
python delete_chara.py --input "待处理.xlsx" --output "处理后.xlsx"

说明：
1) 若不传 --output，默认输出到脚本同目录，文件名为 <输入文件名>_cleaned.xlsx
2) 若 --output 只给文件名（不含目录），也会保存到脚本同目录
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook


# 需要删除的符号（按你提供的截图整理）+ 数字 0-9
# 其中 "..." 作为一个整体也会被移除；其余按单字符移除。
SYMBOLS_TO_REMOVE = [
    "~",
    "，",
    "。",
    "\\",
    ";",
    "?",
    "P",
    "C",
    "_",
    "R",
    "D",
    "•",
    ":",
    "、",
    "!",
    "...",
    "I",
    "N",
    "F",
    "/",
    "O",
    "*",
    "@",
    "【",
    "】",
    "`",
    "〈",
    "〉",
    "“",
    "”",
    "|",
    "；",
    "S",
    "s",
    "-",
    "？",
    "……",
    "：",
    "！",
    "甲組簡",
    "乙組簡",
    "○",
    "%",
]


def build_remove_pattern() -> re.Pattern[str]:
    """构建删除模式：先匹配 ...，再匹配单字符和数字。"""
    single_chars = "".join(ch for ch in SYMBOLS_TO_REMOVE if ch != "...")
    char_class = re.escape(single_chars) + "0-9"
    pattern = r"\.\.\.|[" + char_class + r"]"
    return re.compile(pattern)


REMOVE_PATTERN = build_remove_pattern()


def clean_text(value: str) -> str:
    """先将 = 替换为 #，再删除目标符号和数字。"""
    replaced = value.replace("=", "#")
    return REMOVE_PATTERN.sub("", replaced)


def clean_workbook(input_file: Path, output_file: Path) -> tuple[int, int]:
    """
    清洗整个工作簿（所有工作表、所有字符串单元格）。

    返回：
        (修改单元格数量, 扫描的字符串单元格数量)
    """
    wb = load_workbook(input_file)
    changed_cells = 0
    scanned_text_cells = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value:
                    scanned_text_cells += 1
                    cleaned = clean_text(value)
                    if cleaned != value:
                        cell.value = cleaned
                        changed_cells += 1

    wb.save(output_file)
    return changed_cells, scanned_text_cells


def resolve_output_path(script_dir: Path, input_file: Path, output_arg: str | None) -> Path:
    """输出路径规则：默认或相对路径都落在脚本目录。"""
    if output_arg is None:
        return script_dir / f"{input_file.stem}_cleaned{input_file.suffix}"

    output_path = Path(output_arg)
    if output_path.is_absolute():
        return output_path
    return script_dir / output_path.name


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="删除 Excel 中指定符号并保存为新文件")
    parser.add_argument(
        "--input",
        required=True,
        help="输入 Excel 文件路径（.xlsx）",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="输出文件名或路径；不填则默认 <输入名>_cleaned.xlsx（保存到脚本目录）",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    script_dir = Path(__file__).resolve().parent
    input_file = Path(args.input).resolve()
    output_file = resolve_output_path(script_dir, input_file, args.output)

    if not input_file.exists():
        raise FileNotFoundError(f"找不到输入文件：{input_file}")
    if input_file.suffix.lower() != ".xlsx":
        raise ValueError("仅支持 .xlsx 文件")

    changed_cells, scanned_text_cells = clean_workbook(input_file, output_file)

    print("处理完成")
    print(f"输入文件：{input_file}")
    print(f"输出文件：{output_file}")
    print(f"扫描字符串单元格：{scanned_text_cells}")
    print(f"修改单元格数量：{changed_cells}")
    print("已处理：先将 = 替换为 #，再删除截图符号 + 数字 0-9")


if __name__ == "__main__":
    main()
