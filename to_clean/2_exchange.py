"""
对 delete_chara 脚本产出的 Excel 做二次精细转换。

当前规则：
1) 将 □ 替换为 #
2) 处理括号内容（支持全角（）与半角()）：
   - 括号内为空：删除整段括号
   - 括号内仅 1 个中文：用该中文替换括号前 1 个字符，并删除括号段
   - 括号内长度 >= 2：删除整段括号
   启动方式：python 2_exchange.py --input "待处理.xlsx" --output "处理后.xlsx"
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook


# 匹配全角（）和半角()，且不处理嵌套括号
PAREN_PATTERN = re.compile(r"（([^（）]*)）|\(([^()]*)\)")


def is_single_chinese_char(text: str) -> bool:
    """判断是否恰好为 1 个中文字符。"""
    return len(text) == 1 and "\u4e00" <= text <= "\u9fff"


def transform_parentheses(text: str) -> str:
    """按规则处理括号包裹内容。"""
    result = text
    while True:
        match = PAREN_PATTERN.search(result)
        if not match:
            break

        start, end = match.span()
        inner = match.group(1) if match.group(1) is not None else match.group(2)

        # 规则 A：空括号，直接删除
        if inner == "":
            result = result[:start] + result[end:]
            continue

        # 规则 B：单个中文，替换前一字符后删除括号段
        if is_single_chinese_char(inner):
            if start > 0:
                result = result[: start - 1] + inner + result[end:]
            else:
                # 没有前一个字符可替换时，仅删除括号段
                result = result[:start] + result[end:]
            continue

        # 规则 C：长度 >= 2（或其它未覆盖情形），删除整段括号
        result = result[:start] + result[end:]

    return result


def replace_matched_pairs(
    text: str,
    open_char: str,
    close_char: str,
    handler: Callable[[str], str],
) -> str:
    """
    对成对括号执行替换，使用栈做严密匹配。
    只处理匹配成功的括号对，未匹配字符保持原样。
    """
    result = text
    while True:
        stack: list[int] = []
        pairs: list[tuple[int, int]] = []

        for idx, ch in enumerate(result):
            if ch == open_char:
                stack.append(idx)
            elif ch == close_char and stack:
                start = stack.pop()
                pairs.append((start, idx))

        if not pairs:
            break

        # 从后往前替换，避免索引偏移
        for start, end in reversed(pairs):
            inner = result[start + 1 : end]
            replaced = handler(inner)
            result = result[:start] + replaced + result[end + 1 :]

    return result


def handle_square_brackets(_: str) -> str:
    """
    规则3：[] 包裹的字符串整体替换为 #。
    """
    return "#"


def handle_special_square_brackets(inner: str) -> str:
    """
    规则4：〔〕 包裹的字符串处理。
    - 长度为1：连同该字和括号一起去掉
    - 长度>=2：仅去掉外层括号，保留内容
    - 长度为0：结果为空
    """
    if len(inner) == 1:
        return ""
    if len(inner) >= 2:
        return inner
    return ""


def transform_text(text: str) -> str:
    """先替换 □ 为 #，再执行各类括号规则。"""
    replaced = text.replace("□", "#")
    after_parentheses = transform_parentheses(replaced)
    after_square = replace_matched_pairs(after_parentheses, "[", "]", handle_square_brackets)
    after_special_square = replace_matched_pairs(after_square, "〔", "〕", handle_special_square_brackets)
    return after_special_square


def process_workbook(input_file: Path, output_file: Path) -> tuple[int, int]:
    """
    处理整个工作簿（所有工作表、所有字符串单元格）。

    返回：
        (修改单元格数量, 扫描的字符串单元格数量)
    """
    wb = load_workbook(input_file)
    changed_cells = 0
    scanned_text_cells = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value:
                    scanned_text_cells += 1
                    transformed = transform_text(value)
                    if transformed != value:
                        cell.value = transformed
                        changed_cells += 1

    wb.save(output_file)
    return changed_cells, scanned_text_cells


def resolve_output_path(script_dir: Path, input_file: Path, output_arg: str | None) -> Path:
    """输出路径规则：默认或相对路径都保存到脚本目录。"""
    if output_arg is None:
        return script_dir / f"{input_file.stem}_exchange{input_file.suffix}"

    output_path = Path(output_arg)
    if output_path.is_absolute():
        return output_path
    return script_dir / output_path.name


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Excel 二次转换脚本（exchange）")
    parser.add_argument(
        "--input",
        required=True,
        help="输入 Excel 文件路径（建议为 delete_chara 产出文件）",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="输出文件名或路径；不填则默认 <输入名>_exchange.xlsx（保存到脚本目录）",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    script_dir = Path(__file__).resolve().parent
    input_file = Path(args.input).resolve()
    output_file = resolve_output_path(script_dir, input_file, args.output)

    if not input_file.exists():
        raise FileNotFoundError(f"找不到输入文件：{input_file}")
    if input_file.suffix.lower() != ".xlsx":
        raise ValueError("仅支持 .xlsx 文件")

    changed_cells, scanned_text_cells = process_workbook(input_file, output_file)

    print("处理完成")
    print(f"输入文件：{input_file}")
    print(f"输出文件：{output_file}")
    print(f"扫描字符串单元格：{scanned_text_cells}")
    print(f"修改单元格数量：{changed_cells}")
    print("规则已执行：□ -> #；（）规则；[] -> #；〔〕规则")


if __name__ == "__main__":
    main()
